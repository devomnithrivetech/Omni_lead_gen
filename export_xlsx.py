"""Export leads to Excel with deduplication and clean columns."""
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "leads.db"

# Columns to export (in order)
EXPORT_COLUMNS = [
    ("company_name", "Company Name"),
    ("company_website", "Company Website"),
    ("company_description", "Company Description"),
    ("company_industry", "Industry"),
    ("job_title", "Hiring For (Job Title)"),
    ("job_location", "Location"),
    ("decision_maker_name", "Decision Maker"),
    ("decision_maker_title", "DM Title"),
    ("decision_maker_email", "DM Email"),
    ("decision_maker_linkedin", "DM LinkedIn"),
]


def export():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Get all enriched leads with decision maker email
    rows = conn.execute(
        """SELECT * FROM leads
           WHERE decision_maker_email IS NOT NULL
           AND decision_maker_email != ''
           ORDER BY company_name, created_at"""
    ).fetchall()

    if not rows:
        print("No enriched leads to export.")
        conn.close()
        return

    # Deduplicate: keep only ONE row per company (the one with most data)
    seen = {}
    for row in rows:
        company_key = (row["company_name"] or "").lower().strip()
        if company_key not in seen:
            seen[company_key] = dict(row)
        else:
            # Keep the one with more filled fields
            existing = seen[company_key]
            new = dict(row)
            existing_filled = sum(1 for k, _ in EXPORT_COLUMNS if existing.get(k))
            new_filled = sum(1 for k, _ in EXPORT_COLUMNS if new.get(k))
            if new_filled > existing_filled:
                seen[company_key] = new

    unique_leads = list(seen.values())
    print("Exporting " + str(len(unique_leads)) + " unique companies (from " + str(len(rows)) + " total leads)")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(size=10, name="Arial")
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    # Header row
    for col_idx, (_, header) in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, lead in enumerate(unique_leads, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, 1):
            val = lead.get(key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

    # Column widths
    col_widths = {
        "Company Name": 25,
        "Company Website": 28,
        "Company Description": 45,
        "Industry": 18,
        "Hiring For (Job Title)": 30,
        "Location": 22,
        "Decision Maker": 22,
        "DM Title": 25,
        "DM Email": 30,
        "DM LinkedIn": 35,
    }
    for col_idx, (_, header) in enumerate(EXPORT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 20)

    # Freeze header + auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Row height for header
    ws.row_dimensions[1].height = 30

    outfile = "leads_export.xlsx"
    wb.save(outfile)
    print("Saved: " + outfile)
    conn.close()


if __name__ == "__main__":
    export()